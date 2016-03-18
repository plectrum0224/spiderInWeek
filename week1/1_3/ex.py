#!/usr/bin/env python
# encoding: utf-8


"""
@version: ??
@author: phpergao
@license: Apache Licence 
@contact: endoffight@gmail.com
@site: http://
@software: PyCharm
@file: ex.py
@time: 2016/3/17 21:11
"""
import openpyxl

'''
抓取标题title，地址address，日租金price，第一张房源图片链接houseImg， 房东图片链接hostImg， 房东性别sex，房东名字name
'''

from bs4 import BeautifulSoup
import requests
import time


# 获取单个房源数据函数
def get_webData(url):
    web_data = requests.get(url)
    time.sleep(2)
    soup = BeautifulSoup(web_data.text, 'lxml')

    # 获取对应网页数据
    title = soup.select('body > div.wrap.clearfix.con_bg > div.con_l > div.pho_info > h4 > em')
    address = soup.select('body > div.wrap.clearfix.con_bg > div.con_l > div.pho_info > p > span.pr5')
    price = soup.select('#pricePart > div.day_l > span')
    houseImg = soup.select('#curBigImage')
    hostImg = soup.select('div.js_box.clearfix > div.member_pic > a > img')
    sex = soup.select('div.js_box.clearfix > div.member_pic > div')
    name = soup.select('div.js_box.clearfix > div.w_240 > h6 > a')

    # 提取具体信息
    for title_,address_,price_,houseImg_,hostImg_,sex_,name_ in zip(title, address, price, houseImg, hostImg, sex, name):
        data = {
            'title': title_.get_text(),
            'address': address_.get_text().strip(),
            'price': price_.get_text(),
            'houseImg': houseImg_.get('src'),
            'hostImg': hostImg_.get('src'),
            'sex': sex_.get('class'),
            'name': name_.get_text(),
        }

        # 判断房东性别
        if data['sex'][0]=='member_ico':
            data['sex'] = 'male'
        else:
            data['sex'] = 'female'

        return data

#获取多页的房屋信息
def get_infos(page):
    print('Ready....')
    count = 1
    item = 1
    urls = ['http://bj.xiaozhu.com/search-duanzufang-p{}-0/'.format(str(i)) for i in range(1, page+1)]
    for url in urls:
        web_links = requests.get(url)
        time.sleep(2)
        soup = BeautifulSoup(web_links.text, 'lxml')

        links = soup.select('#page_list > ul > li > a')

        for link in links:
            pageLink = (link.get('href'))
            dict_ = get_webData(pageLink)
            print('writing...')
            writeToExcel(item, dict_) #写入数据到excel
            print('itme{} finished...'.format(item))
            item += 1
        print('page{} finished{}'.format(count, '-'*30))
        count += 1




#新建一个Excel表格
def initExcel():
    wb = openpyxl.Workbook() #新建excel文件
    wb.create_sheet(index=0, title='houseInfo') #新建一个sheet，名称为houseInfo
    sheet = wb.get_sheet_by_name('houseInfo')
    headText = ['Title', 'Adress', 'Price', 'Sex', 'Name', 'HouseImg', 'HostImg']
    for columnNum in range(1, len(headText)+1):
        sheet.cell(row=1, column=columnNum).value = headText[columnNum-1] #将标题写到excel中
    wb.save('House Info.xlsx')

#写入数据到Excel
def writeToExcel(rowNum, dict_=None):
    wb = openpyxl.load_workbook('House Info.xlsx')
    sheet = wb.get_sheet_by_name('houseInfo')
    try:
        sheet.cell(row = rowNum+1, column=1).value = dict_['title']
        sheet.cell(row = rowNum+1, column=2).value = dict_['address']
        sheet.cell(row = rowNum+1, column=3).value = dict_['price']
        sheet.cell(row = rowNum+1, column=4).value = dict_['sex']
        sheet.cell(row = rowNum+1, column=5).value = dict_['name']
        sheet.cell(row = rowNum+1, column=6).value = dict_['houseImg']
        sheet.cell(row = rowNum+1, column=7).value = dict_['hostImg']
        wb.save('House Info.xlsx')
    except TypeError:
        print('InValid Info')


if __name__ == '__main__':
    initExcel()
    get_infos(10)



